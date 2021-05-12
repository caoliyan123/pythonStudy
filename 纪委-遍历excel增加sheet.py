#coding:utf-8

import os
import pandas as pd
from openpyxl import load_workbook
import xlrd,xlwt

sheetname=['基本情况','出国证件','个人简历','家庭成员','奖励情况','年度考核','兼职情况','房产情况','股票','基金','投资型保险','经商办企业','国境外存款','婚姻情况','出国情况','外国人通婚','移居国外','刑事情况','国境外投资','从业情况','金融理财','工资情况','劳务所得','用车情况','第一形态','车辆情况','境内主管','求学情况','婚丧喜庆']

try:
#遍历所有文件夹及子文件夹
    for root_dir,sub_dir,files in os.walk('C:\\Users\\Yan\\Desktop\\子表4'):
        for f in files:
            if f.endswith('.xlsx') :
                #获取绝对路径
                file_name=os.path.join(root_dir,f)
                print (file_name)
                #打开excel并读取所有sheet
                sheet = pd.read_excel(file_name, sheet_name=None)
                #将sheet转换成list列表
                sheet=list(sheet)
                book=load_workbook(file_name)
                wirter=pd.ExcelWriter(file_name, engine='openpyxl')
                wirter.book=book
                df=pd.DataFrame()
                #对比两个列表中sheet名称,将不存在于sheet的新增
                for s in sheetname:
                    if s not in sheet:
                        df.to_excel(wirter,sheet_name=s)
                wirter.save()
            elif (f.endswith('.xls')):
                # 获取绝对路径
                file_name = os.path.join(root_dir, f)
                print(file_name)
                # 打开excel并读取所有sheet
                sheet = pd.read_excel(file_name, sheet_name=None)
                # 将sheet转换成list列表
                sheet = list(sheet)
                book = load_workbook(file_name)
                wirter = pd.ExcelWriter(file_name)
                wirter.book = book
                df = pd.DataFrame()
                # 对比两个列表中sheet名称,将不存在于sheet的新增
                for s in sheetname:
                    if s not in sheet:
                        df.to_excel(wirter, sheet_name=s)
                wirter.save()

except Exception  as e:
    print(e)




