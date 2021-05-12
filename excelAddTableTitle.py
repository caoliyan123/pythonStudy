#coding:utf-8

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook

#房产情况
house=['产权人','称谓','共有产权情况','房产来源（去向）','房产坐落地址','建筑时间','建筑面积(m²)','房产性质','交易时间','资金来源','交易金额','备注']
#股票
stock=['姓名','称谓','股票名称或代码','持股数量','填报前一交易日市值','记录时间']

#基金
fund=['姓名','称谓','基金名称或代码','基金份额','填报前一交易日净值（万元）','记录时间']

#投资型保险
insurance=['投保人姓名','称谓','保险产品全称及保单号','保险公司名称','累积缴纳保费、投资金（万元）','记录时间']

#经商办企业
business=['姓名','称谓','是否为境外企业','企业或其他市场主体名称','职务','经营范围','统一社会信用代码/注册号','成立日期','注册地','经营地','企业或其他市场主体类型','注册资本（金）或资金数额（出资额）（万元）','出资金额个人认缴出资额或个人出资额（万元）','个人认缴出资比例或个人出资比例（%）','备注']


#国境外存款
out_deposit=['统一社会信用代码/注册号','成立日期','注册地','经营地','企业或其他市场主体类型','注册资本（金）或资金数额（出资额）（万元）','出资金额个人认缴出资额或个人出资额（万元）','个人认缴出资比例或个人出资比例（%）','备注']

#婚姻情况
marry=['当前婚姻状况','本年度婚姻变化','变化状态','变化时间']

#出国（境）情况
abroad=['起始时间','截止时间','所到国家(地区)','出国(境)事由','审批机构','经费来源','所用护照(证件)号码']

#外国人通婚
foreign_marriage=['姓名','称谓','配偶姓名','国籍（地区）','工作（学习）单位','职务','登记时间','婚后居住地']

#移居国外
emigrate=['姓名','称谓','移居国家（地区）','现居住地','移居证件号码','移居类别','移居时间','备注']

#刑事情况
criminal=['姓名','称谓','被追究的时间','被追究的原因','处理机关','处理阶段','处理结果']

#国境外投资
overseas_bussiness=['投资人姓名','称谓','投资或入股单位','投资的国家（地区）及城市','投资或入股方式','担任职务','投资情况','投资币种','出资金额','人名币金额','年收益','备注']

#从业情况
work=['姓名','称谓','是否共同生活','证件类型','证件号码','单位名称','单位性质','是否担任高级职务','职务','备注']

#金融理财
financial=['姓名','称谓','产品名称','持有份额','市值（万元）','记录时间']

#工资情况
salary=['工资（含津贴、补贴）（万元/全年）','兼职收入','奖金（万元/全年）','其他（万元/全年）','合计（万元/全年）']

#劳务所得
income=['讲学(万元/全年)','写作(万元/全年)','咨询(万元/全年)','审稿(万元/全年)','书画(万元/全年)','其他(万元/全年)','合计(万元/全年)']

#用车情况
usecar=['办公用房','有无用车情况','价格（万元）','排气量（升）']

#第一形态
first=['时间','实施人（部门）','原因','内容','本人整改（回复）情况','处理方式']

#车辆情况
car=['车主','购车时间','购车价格','品牌型号','车辆号牌']

#境内主管
director=['姓名','称谓','是否共同生活','证件类型','证件号码','企业名称','所属国家地区','分支机构地址','所任职务','备注']

#求学情况
school=['子女姓名','称谓','出国时间','求学国家、地区市','学校名称','每年留学费用（万元）','费用来源','备注']

#婚丧喜庆
Mar_fun_jub=['与本人关系','操办事项','办理时间','办理地点','参加对象及人数','接受礼品及处理情况','有否报告']


def addTitle(dirpath):
    filecount=0
    errorcount=0
    for root_dir, sub_dir, files in os.walk(dirPath):
        for f in files:
            try:
                if (f.endswith('.xlsx') ):
                    # 获取绝对路径
                    file_name = os.path.join(root_dir, f)
                    print(file_name)
                    excel=openpyxl.load_workbook(file_name)
                    # 房产
                    sheet_1=excel['房产情况']
                    #sheet_1.append(house)
                    if sheet_1.cell(row=1,column=1).value is None:
                        for i,j in zip(range(len(house)),house):
                            sheet_1.cell(row=1,column=i+1,value=j)
                            i+=1

                    #股票
                    sheet_2=excel['股票']
                    if sheet_2.cell(row=1,column=1).value is None:
                        for i, j in zip(range(len(stock)), stock):
                            sheet_2.cell(row=1, column=i + 1, value=j)
                            i += 1
                     #基金
                    sheet_3 = excel['基金']
                    if sheet_3.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(fund)), fund):
                            sheet_3.cell(row=1, column=i + 1, value=j)
                            i += 1
                    #投资型保险
                    sheet_4 = excel['投资型保险']
                    if sheet_4.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(insurance)), insurance):
                            sheet_4.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 经商办企业 business
                    sheet_5 = excel['经商办企业']
                    if sheet_5.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(business)), business):
                            sheet_5.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 国境外存款 out_deposit
                    sheet_6 = excel['国境外存款']
                    if sheet_6.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(out_deposit)), out_deposit):
                            sheet_6.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 婚姻情况 marry
                    sheet_7 = excel['婚姻情况']
                    if sheet_7.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(marry)), marry):
                            sheet_7.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 出国（境）情况 abroad
                    sheet_8 = excel['出国（境）情况']
                    if sheet_8.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(abroad)), abroad):
                            sheet_8.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 外国人通婚 foreign_marriage
                    sheet_9 = excel['外国人通婚']
                    if sheet_9.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(foreign_marriage)), foreign_marriage):
                            sheet_9.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 移居国外 emigrate
                    sheet_10 = excel['移居国外']
                    if sheet_10.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(emigrate)), emigrate):
                            sheet_10.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 刑事情况 criminal
                    sheet_11 = excel['刑事情况']
                    if sheet_11.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(criminal)), criminal):
                            sheet_11.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 国境外投资 overseas_bussiness
                    sheet_12 = excel['国境外投资']
                    if sheet_12.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(overseas_bussiness)), overseas_bussiness):
                            sheet_12.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 从业情况 work
                    sheet_13 = excel['从业情况']
                    if sheet_13.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(work)), work):
                            sheet_13.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 金融理财 financial
                    sheet_14 = excel['金融理财']
                    if sheet_14.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(financial)), financial):
                            sheet_14.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 工资情况 salary
                    sheet_15 = excel['工资情况']
                    if sheet_15.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(salary)), salary):
                            sheet_15.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 劳务所得 income
                    sheet_16 = excel['劳务所得']
                    if sheet_16.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(income)), income):
                            sheet_16.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 用车情况 usecar
                    sheet_17 = excel['用车情况']
                    if sheet_17.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(usecar)), usecar):
                            sheet_17.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 第一形态 first
                    sheet_18 = excel['第一形态']
                    if sheet_18.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(first)), first):
                            sheet_18.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 车辆情况 car
                    sheet_19 = excel['车辆情况']
                    if sheet_19.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(car)), car):
                            sheet_19.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 境内主管 director
                    sheet_20 = excel['境内主管']
                    if sheet_20.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(director)), director):
                            sheet_20.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 求学情况 school
                    sheet_21 = excel['求学情况']
                    if sheet_21.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(school)), school):
                            sheet_21.cell(row=1, column=i + 1, value=j)
                            i += 1
                    # 婚丧喜庆 Mar_fun_jub
                    sheet_22 = excel['婚丧喜庆']
                    if sheet_22.cell(row=1, column=1).value is None:
                        for i, j in zip(range(len(Mar_fun_jub)), Mar_fun_jub):
                            sheet_22.cell(row=1, column=i + 1, value=j)
                            i += 1
                    excel.save(file_name)
                    filecount+=1
            except Exception as e:
                print(file_name,'------文件报错------' ,e)
                errorcount+=1
                pass
            continue
    return '已处理 {0} 个文件,出错文件 {1} 个'.format(filecount, errorcount)
if __name__ == '__main__':
    dirPath = r'C:\Users\Yan\Desktop\2021-5-12\1'
    print(addTitle(dirPath))
