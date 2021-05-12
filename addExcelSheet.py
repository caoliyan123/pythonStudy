#coding:utf-8

import os
import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32

sheetname=['基本情况','出入国（境）证件','个人简历','家庭成员和重要社会关系','奖励情况','年度考核','兼职情况','房产情况','股票','基金','投资型保险','经商办企业','国境外存款','婚姻情况','出国（境）情况','外国人通婚','移居国外','刑事情况','国境外投资','从业情况','金融理财','工资情况','劳务所得','用车情况','第一形态','车辆情况','境内主管','求学情况','婚丧喜庆']

def xlstoxlsx(f):
    fname=dirPath
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)
    wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    n_fname=fname.split('.')[0]+'.xlsx'
    os.remove(fname)
    return n_fname

def addSheet(dirPath):
    filecount=0
    errorfile=0
    # 遍历所有文件夹及子文件夹
    try:
        for root_dir, sub_dir, files in os.walk(dirPath):
            for f in files:
                try:
                    if (f.endswith('.xlsx') ):
                        # 获取绝对路径
                        file_name = os.path.join(root_dir, f)
                        #print(file_name)
                        # 打开excel并读取所有sheet
                        sheet = pd.read_excel(file_name, sheet_name=None)
                        # 将sheet转换成list列表
                        sheet = list(sheet)
                        #print(sheet)
                        book = load_workbook(file_name)
                        wirter = pd.ExcelWriter(file_name, engine='openpyxl')
                        wirter.book = book
                        df = pd.DataFrame()
                        # 对比两个列表中sheet名称,将不存在于sheet的新增
                        for s in sheetname:
                            if s not in sheet:
                                df.to_excel(wirter, sheet_name=s)
                        wirter.save()
                        filecount+=1
                    elif (f.endswith('.xls') ):
                        # 获取绝对路径
                        file_name = os.path.join(xlstoxlsx(f))
                        # print(file_name)
                        # 打开excel并读取所有sheet
                        sheet = pd.read_excel(file_name, sheet_name=None)
                        # 将sheet转换成list列表
                        sheet = list(sheet)
                        print(sheet)
                        book = load_workbook(file_name)
                        wirter = pd.ExcelWriter(file_name, engine='openpyxl')
                        wirter.book = book
                        df = pd.DataFrame()
                        # 对比两个列表中sheet名称,将不存在于sheet的新增
                        for s in sheetname:
                            if s not in sheet:
                                df.to_excel(wirter, sheet_name=s)
                        wirter.save()
                        filecount += 1
                except Exception as e:
                    print('发生异常,文件处理出错',file_name)
                    errorfile+=1
                    pass
                continue

        print ("已处理{0}个正确文件,{1}个文件处理出错".format(filecount,errorfile))
    except OSError as e:
        print('目录不存在',e)

if __name__ == '__main__':
    dirPath = r'C:\Users\Yan\Desktop\2021-5-12\1'
    print(addSheet(dirPath))
    #print(addSheet(dirPath))
